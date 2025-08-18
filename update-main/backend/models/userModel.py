# import bcrypt
from db import getConnection
from openpyxl import load_workbook
import random


        # DROPDOWN NAVBAR FRONTEND (KATEGORI DAN SUBKATEGORI)
def listKategori():
    conn = getConnection()
    ws = conn['Kategori']
    listKategori = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        idKategori = row[0]
        kategori = row[1]

        listKategori.append({
            'id': idKategori,
            'kategori': kategori
        })
    return listKategori


def listSubKategori():
    conn = getConnection()
    ws = conn['Subkategori']
    listSubKategori = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        idSubKategori = row[0]
        idKategori = row[1]
        subKategoriName = row[2]

        if idSubKategori and idKategori:
            if idKategori not in listSubKategori:
                listSubKategori[idKategori] = {
                    'idKategori': idKategori,
                    'subKategori': []
                }

            if subKategoriName:
                listSubKategori[idKategori]['subKategori'].append({
                    'id': idSubKategori,
                    'subKategoriName': subKategoriName
                })

    return list(listSubKategori.values())



        # SEMUA PRODUK
def allProduct():
    conn = getConnection()
    ws = conn['Produk']
    allProduct = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_product = row[0]
        name_product = row[1]
        id_kategori = row[2]
        id_subkategori = row[3]
        deskripsi = row[4]
        harga = row[5]
        stok = row[6]
        berat = row[7]
        rating = row[8]
        promo = row[9]

        if id_product: 
            product_data = {
                "id": id_product,
                "nama_produk": name_product,
                "id_kategori": id_kategori,
                "id_subkategori": id_subkategori,
                "deskripsi": deskripsi,
                "harga": harga,
                "stok": stok,
                "berat": berat,
                "rating": rating,
                "promo": promo
            }
            allProduct.append(product_data)

    result = {}
    for p in allProduct:
        kategori = str(p["id_kategori"])
        subkategori = str(p["id_subkategori"])
        
        if kategori not in result:
            result[kategori] = {}
        if subkategori not in result[kategori]:
            result[kategori][subkategori] = []
        
        result[kategori][subkategori].append(p)

    return result



        # PRODUK REKOMENDASI
def listProdukRekomendasi():
    conn = getConnection()
    ws = conn['Produk_Rekomendasi']
    listProdukkRekomendasi = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        id_rekomendasi = row[0]
        id_produk = row[1]
        id_produk_rekomendasi = row[2]

        if id_rekomendasi and id_produk:
            if id_produk not in listProdukkRekomendasi:
                listProdukkRekomendasi[id_produk] = {
                    'id_produk': id_produk,
                    'list_rekomendasi': []
                }

            if id_produk_rekomendasi:
                listProdukkRekomendasi[id_produk]['list_rekomendasi'].append({
                    'id': id_rekomendasi,
                    'id_produk_rekomendasi': id_produk_rekomendasi
                })

    return list(listProdukkRekomendasi.values())


# def getAllKategori():
#     conn = getConnection()
#     cur = conn.cursor(dictionary=True)
#     try :    
#         cur.execute("SELECT * FROM Kategori")
#         result = cur.fetchall()
#         return result
#     except Exception as e:
#         print("gagal mengambil data dari database", e)
#         return None
#     finally:
#         if conn:
#             conn.close()

# def getAllKategori():
#     try:
#         df = getConnection(sheet_name="Kategori") 
#         return df.to_dict(orient="records") 
#     except Exception as e:
#         print("Gagal mengambil data dari Excel:", e)
#         return None

#

            
            # ====== LOGIN ======
# def loginUser(username, password):
#     conn = getConnection()
#     cur = conn.cursor(dictionary=True)
#     try:
#         cur.execute("SELECT * FROM users WHERE username=%s", (username,))
#         user = cur.fetchone()
#         if user and bcrypt.checkpw(password.encode('utf-8'), user['password'].encode('utf-8')):
#             return user
#         else:
#             return None
#     except Exception as e:
#         print("Error saat login:", e)
#         return None
#     finally:
#         if conn:
#             conn.close()

# def registerUser(username, email, password):
#     conn = getConnection()
#     cur = conn.cursor()
#     try:
#         hashed = bcrypt.hashpw(password_plain.encode('utf-8'), bcrypt.gensalt())
#         cur.execute("INSERT INTO users (username, email, password) VALUES (%s, %s, %s)", 
#                     (username, email, hashed.decode('utf-8')))
#         conn.commit()
#         return True
#     except Exception as e:
#         print("Error saat register:", e)
#         return False
#     finally:
#         if conn:
#             conn.close()




def addUser(name, user):
    conn = getConnection()
    cur = conn.cursor()
    try :
        cur.execute("INSERT INTO users (name, user) VALUES (%s, %s)", (name, user))
        conn.commit()
        return True
    except Exception as e:
        print("Gagal menambahkan user:", e)
        return False
    finally:
        if conn:
            conn.close()


def updateUser(id, name, user):
    conn = getConnection()
    cur = conn.cursor()
    try:    
        cur.execute("UPDATE users SET name=%s, user=%s WHERE id=%s", (name, user, id))
        conn.commit()
        return True
    except Exception as e:
        print("gagal memperbarui user", e)
        return False
    finally:
        if conn:
            conn.close()

def deleteUser(id):
    conn = getConnection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM users WHERE id=%s", (id,))
        conn.commit()
        return True
    except Exception as e:
        print("gagal hapus user", e) 
        return False
    finally:
        if conn:
            conn.close()

